from __future__ import print_function

import datetime
import os.path

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

from openpyxl import load_workbook

class work_event:
    def __init__(self, start, end) -> None:
        sep_T = 'T'
        sep_plus = '+'
        sep_colon = ':'
        sep_hyphen = '-'
        start_date, _sep, start = start.partition(sep_T)
        end_date, _sep, end = end.partition(sep_T)

        self.start_year, _sep, start_date = start_date.partition(sep_hyphen)
        self.end_year, _sep, end_date = end_date.partition(sep_hyphen)

        self.start_month, _sep, self.start_day = start_date.partition(sep_hyphen)
        self.end_month, _sep, self.end_day = end_date.partition(sep_hyphen)
        
        start, _sep, _after = start.partition(sep_plus)
        end, _sep, _after = end.partition(sep_plus)
        
        self.start_hour, _sep, start = start.partition(sep_colon)
        self.end_hour, _sep, end = end.partition(sep_colon)
        
        self.start_minute, _sep, self.start_second = start.partition(sep_colon)
        self.end_minute, _sep, self.end_second = end.partition(sep_colon)

# If modifying these scopes, delete the file token.json.
SCOPES = ['https://www.googleapis.com/auth/calendar.readonly']

def main():
    """Shows basic usage of the Google Calendar API.
    Prints the start and name of the next 10 events on the user's calendar.
    """
    creds = None
    creds_filename = 'a.json'
    # The file token.json stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists(creds_filename):
        creds = Credentials.from_authorized_user_file(creds_filename, SCOPES)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open('token.json', 'w') as token:
            token.write(creds.to_json())

    try:
        service = build('calendar', 'v3', credentials=creds)

        # Call the Calendar API
        event_cout = 900
        print('Getting the upcoming ' + str(event_cout) + ' events')
        events_result = service.events().list(calendarId='primary', timeMin='2023-03-01T01:00:00Z',
                                              maxResults=event_cout, singleEvents=True,
                                              orderBy='startTime', q='Praca').execute()
        events = events_result.get('items', [])

        if not events:
            print('No upcoming events found.')
            return

        work_events = []
        for event in events:
            start = event['start'].get('dateTime', event['start'].get('date'))
            end = event['end'].get('dateTime', event['end'].get('date'))
            new_work_event = work_event(start, end)
            work_events.append(new_work_event)

        filename = "working_time/working_time_aberg.xlsx"
        workbook = load_workbook(filename=filename)
        sheet = workbook.active
        column_offset = 1
        row_offset = 2
        year_offset_23 = -2
        year_offset_24 = 10
        days = {}
        new_day = []
        latest_date = ()
        for event in work_events:
            if(latest_date == ()):
                latest_date = (event.start_year, event.start_month, event.start_day)
            actual_date = (event.start_year, event.start_month, event.start_day)
            if(latest_date == actual_date):
                new_day.append(event)
            if(latest_date != actual_date):
                days[latest_date] = new_day
                latest_date = actual_date
                new_day = []
                new_day.append(event)


        for day in days:
            if(day[0] == '2023'):
                cell_to_modify = sheet.cell(row=((row_offset+int(day[1]) + year_offset_23)), column=(column_offset+int(day[2])))
            if(day[0] == '2024'):
                cell_to_modify = sheet.cell(row=((row_offset+int(day[1]) + year_offset_24)), column=(column_offset+int(day[2])))
            time_sum = 0
            for event in days[day]:
                start_time = float(event.start_hour) + (float(event.start_minute) / 60) + (float(event.start_second)/3600)
                end_time = float(event.end_hour) + (float(event.end_minute) / 60) + (float(event.end_second)/3600)
                time_sum += (end_time - start_time)
            
            cell_to_modify.value = time_sum

        workbook.save(filename=filename)


    except HttpError as error:
        print('An error occurred: %s' % error)


if __name__ == '__main__':
    main()